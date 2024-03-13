from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# These are our Libraries Standards to be used in the Project
# Every File to use with this code must be in same directory or its subdirectory

wb = load_workbook("Grades.xlsx")
ws = wb.active
print(ws["A1"].value)

# We can change the value of a cell by ws['A2'].value = "test"
# We can also do this without '.value'

wb.save("Grades.xlsx")

print(wb.sheetnames)
# To access a specific Work Sheet print(wb['Sheet2'])
# To create a new WorkSheet wb.create_sheet("NewSheet")
# If we create a workbook from scratch, we might wanna insert rows and columns
# instead of values in each cell

ws.append(["Marichu", 60, 50, 70, 60])

for row in range(1, 11):
    for col in range(1, 5):
        char = get_column_letter(col)
        # We can alternatively do it Manually: char = chr(65 + col)
        print(ws[char + str(row)].value)

ws.merge_cells("A1:D3")
ws.unmerge_cells("A1:D3")  # Once it's merged, its data is emptied

ws.insert_rows(7)  # inserted after row 7
ws.delete_rows(7)

ws.insert_cols(2)

ws.move_range("C1:D11", rows=2, cols=2)

# Creating a new Worksheet by inserting data
data = {
    "Joe": {"math": 65, "science": 78, "english": 98, "gym": 89},
    "Bill": {"math": 55, "science": 72, "english": 87, "gym": 95},
    "Tim": {"math": 100, "science": 45, "english": 75, "gym": 92},
    "Sally": {"math": 30, "science": 25, "english": 45, "gym": 100},
    "Jane": {"math": 100, "science": 100, "english": 100, "gym": 60},
}

wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ["Name"] + list(data["Joe"].keys())
ws.append(headings)

for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

for col in range(2, len(data["Joe"]) + 2):  # range from 2 to (4 + 2) 6.
    char = get_column_letter(col)
    ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

for col in range(1, 6):
    ws[get_column_letter(col) + "1"].font = Font(bold=True, color="0099CCFF")

wb.save("NewGrades.xlsx")
